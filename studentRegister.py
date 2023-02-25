from tkinter import *
from tkinter.ttk import Combobox
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from datetime import datetime
from PIL import Image, ImageTk
import os
import openpyxl, xlrd
import pathlib
import tkinter as tk

#constants
PATH = 'C:/Users/ferna/Downloads/python/pythonExcel/'
file = pathlib.Path('C:/Users/ferna/Downloads/python/pythonExcel/Student_data.xlsx')
background="#06283D"
framebg="#EDEDED"
framefg="#06283D"

root = Tk()
root.title('Student registration system')
root.geometry('1250x700+210+100')
root.configure(bg=background)
root.resizable(False, False)



if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of registration"
    sheet['G1']="DNI"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's occupation"
    sheet['L1']="Mother's occupation"
    file.save('C:/Users/ferna/Downloads/python/pythonExcel/Student_data.xlsx')



def Exit():
    root.destroy()


def Show_image():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(), title="Select imagge file",
                                         filetypes=(("JPG File","*.jpg"),("PNG File","*.png"),("ALL files","*.txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2



def Registration_no():
    file=openpyxl.load_workbook(PATH+'Student_data.xlsx') 
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row, column=1).value

    try:
        registration.set(max_row_value+1)

    except:
        registration.set("1")


#clear
def Clear():
    name.set('')
    date_form.set('')
    DNI.set('')
    skill.set('')
    f_name.set('')
    m_name.set('')
    father_occupation.set('')
    mother_occupation.set('')
    classes.set('Select Class')

    saveButton.config(state='normal')

    img1=PhotoImage(file=PATH+"Images/AccountUser2.png")
    lbl.config(image=img1)
    lbl.image=img1
    img=""
    Registration_no()



def Save():
    R1=registration.get()
    N1=name.get()
    C1=classes.get()

    try:
        G1=radio.get()
    except:
        messagebox.showerror("error","Select Gender!")
    
    D1=date.get()
    D2=date_form.get()
    dni1=DNI.get()
    S1=skill.get()
    fatherName=f_name.get()
    motherName=m_entry.get()
    F1=father_occupation.get()
    M1=mother_occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or R1=="" or S1=="" or  dni1=="" or fatherName=="" or motherName=="":
        messagebox.showerror("Error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook(PATH+'Student_data.xlsx')
    sheet=file.active

    sheet.cell(column=1, row=sheet.max_row+1, value=R1)
    sheet.cell(column=2, row=sheet.max_row, value=N1)
    sheet.cell(column=3, row=sheet.max_row, value=C1)
    sheet.cell(column=4, row=sheet.max_row, value=G1)
    sheet.cell(column=5, row=sheet.max_row, value=D2)
    sheet.cell(column=6, row=sheet.max_row, value=D1)
    sheet.cell(column=7, row=sheet.max_row, value=dni1)
    sheet.cell(column=8, row=sheet.max_row, value=S1)
    sheet.cell(column=9, row=sheet.max_row, value=fatherName)
    sheet.cell(column=10, row=sheet.max_row, value=motherName)
    sheet.cell(column=11, row=sheet.max_row, value=F1)
    sheet.cell(column=12, row=sheet.max_row, value=M1)

    file.save(PATH+'Student_data.xlsx')

    try:
        img.save(PATH+"Images/userImages/"+str(R1)+".jpg")
    except:
        messagebox.showerror("info", "Profile picture is not available")

    messagebox.showinfo("info", "Successfully data entered!")

    Clear()
    Registration_no()


def Search_student():
    text=search.get()
    Clear()
    saveButton.config(state="disable")
    file=openpyxl.load_workbook(PATH+'Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            t_row=row[0]

            reg_no_position=str(t_row)[14:-1]
            reg_number=str(t_row)[15:-1]


    try:
        print(str(t_row))
    except:
        messagebox.showerror("Invalid","Invalid registration number!")


    x1=sheet.cell(row=int(reg_number), column=1).value
    x2=sheet.cell(row=int(reg_number), column=2).value
    x3=sheet.cell(row=int(reg_number), column=3).value
    x4=sheet.cell(row=int(reg_number), column=4).value
    x5=sheet.cell(row=int(reg_number), column=5).value
    x6=sheet.cell(row=int(reg_number), column=6).value
    x7=sheet.cell(row=int(reg_number), column=7).value
    x8=sheet.cell(row=int(reg_number), column=8).value
    x9=sheet.cell(row=int(reg_number), column=9).value
    x10=sheet.cell(row=int(reg_number), column=10).value
    x11=sheet.cell(row=int(reg_number), column=11).value
    x12=sheet.cell(row=int(reg_number), column=12).value


    registration.set(x1)
    name.set(x2)
    classes.set(x3)

    if x4=="Female":
        radio.set(value=2)
    else:
        radio.set(value=1)


    date.set(x5)
    date_form.set(x6)
    DNI.set(x7)
    skill.set(x8)
    f_name.set(x9)
    m_name.set(x10)
    father_occupation.set(x11)
    mother_occupation.set(x12)

    img=(Image.open(PATH+"Images/userImages/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2



def Update():
    R1=registration.get()
    N1=name.get()
    C1=classes.get()
    G1=selection()
    D1=date.get()
    D2=date_form.get()
    dni1=DNI.get()
    S1=skill.get()
    fatherName=f_name.get()
    motherName=m_entry.get()
    F1=father_occupation.get()
    M1=mother_occupation.get()

    file=openpyxl.load_workbook(PATH+'Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(R1):
            t_row=row[0]

            reg_no_position=str(t_row)[14:-1]
            reg_number=str(t_row)[15:-1]

#    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=dni1)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=fatherName)
    sheet.cell(column=10, row=int(reg_number), value=motherName)
    sheet.cell(column=11, row=int(reg_number), value=F1)
    sheet.cell(column=12, row=int(reg_number), value=M1)


    file.save(PATH+'Student_data.xlsx')

    try:
        img.save(PATH+"Images/userImages/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update", "Update Successfully!")

    Clear()
    Registration_no()



def selection():
    value=radio.get()
    if value==1:
        gender="Male"
        return gender
    else:
        gender="Female"
        return gender



#Top frame
Label(root, text="Email: hortafer02@gmail.com", width=10, height=3, bg="#f0687c", anchor="e").pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)




#search box
search=StringVar()
Entry(root, textvariable=search, width=15, bd=2, font="arial 20").place(x=820, y=70)
image_icon3=PhotoImage(file=PATH+"Images/search.png")
search_button=Button(root, text="Search", compound=LEFT, image=image_icon3, width=123, bg="#68ddfa", font="arial 13 bold", command=Search_student)
search_button.place(x=1060, y=56)

image_icon4=PhotoImage(file=PATH+"Images/layer.png")
Update_button=Button(root, image=image_icon4, bg="#c36464", command=Update)
Update_button.place(x=110, y=50)


#Registration and Date
Label(root, text="Registration No: ", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date: ", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

registration=IntVar()
date=StringVar()

reg_entry=Entry(root, textvariable=registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)



Registration_no()


today=datetime.today()
d1= today.strftime("%d/%m/%Y")
date_entry=Entry(root, textvariable=date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

date.set(d1)


#Student details
obj=LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name: ", font="arial 13", bg=framebg, fg= framefg).place(x=30, y=50)
Label(obj, text="Date: ", font="arial 13", bg=framebg, fg= framefg).place(x=30, y=100)
Label(obj, text="Gender: ", font="arial 13", bg=framebg, fg= framefg).place(x=30, y=150)

Label(obj, text="Class: ", font="arial 13", bg=framebg, fg= framefg).place(x=500, y=50)
Label(obj, text="DNI: ", font="arial 13", bg=framebg, fg= framefg).place(x=500, y=100)
Label(obj, text="Skills: ", font="arial 13", bg=framebg, fg= framefg).place(x=500, y=150)

name=StringVar()
name_entry=Entry(obj, textvariable=name, width=20, font="arial 10")
name_entry.place(x=160, y=50)


date_form=StringVar()
date_entry2=Entry(obj, textvariable=date_form, width=20, font="arial 10")
date_entry2.place(x=160, y=100)


radio=IntVar()
radio_button1=Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
radio_button1.place(x=150, y=150)

radio_button2=Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
radio_button2.place(x=200, y=150)


DNI=IntVar()
DNI.set('')
dni_entry=Entry(obj, textvariable=DNI, width=20, font="arial 10")
dni_entry.place(x=630, y=100)


skill=StringVar()
skill_entry=Entry(obj, textvariable=skill, width=20, font="arial 10")
skill_entry.place(x=630, y=150)


classes=Combobox(obj, values=['1','2','3','4','5','6','7','8','9','10','11', '12'], font="Roboto 10", width=17, state="r")
classes.place(x=630, y=50)
classes.set("Select Class")



#Parent's details
obj2=LabelFrame(root, text="Parent's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)


Label(obj2, text="Father's Name: ", font="arial 13", bg=framebg, fg= framefg).place(x=30, y=50)
Label(obj2, text="Occupation: ", font="arial 13", bg=framebg, fg= framefg).place(x=30, y=100)

f_name=StringVar()
f_entry=Entry(obj2, textvariable=f_name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

father_occupation=StringVar()
fo_entry=Entry(obj2, textvariable=father_occupation, width=20, font="arial 10")
fo_entry.place(x=160, y=100)


Label(obj2, text="Mother's Name: ", font="arial 13", bg=framebg, fg= framefg).place(x=500, y=50)
Label(obj2, text="Occupation: ", font="arial 13", bg=framebg, fg= framefg).place(x=500, y=100)

m_name=StringVar()
m_entry=Entry(obj2, textvariable=m_name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

mother_occupation=StringVar()
mo_entry=Entry(obj2, textvariable=mother_occupation, width=20, font="arial 10")
mo_entry.place(x=630, y=100)


#image
f=Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img=PhotoImage(file=PATH+"Images/AccountUser2.png")
lbl=Label(f, bg="black", image=img)
lbl.place(x=0, y=0)


#button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=Show_image).place(x=1000, y=370)
saveButton=Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)


root.mainloop()